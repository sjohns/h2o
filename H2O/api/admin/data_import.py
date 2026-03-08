from __future__ import annotations

import json
from io import BytesIO
from pathlib import Path
from typing import Any, BinaryIO

from openpyxl import load_workbook
from pydantic import BaseModel, ConfigDict, Field, ValidationError, field_validator


SHEET_NAME = "SKUs"
SHEET_NAME_PRODUCT_TYPES = "Product Types"

# Display name → internal field name mappings for the two-sheet export format
_TWO_SHEET_PT_MAP = {
    "Name": "product_type_name",
    "Display Order": "product_type_display_order",
    "Active": "product_type_active_flag",
}
_TWO_SHEET_SKU_MAP = {
    "Product Type": "product_type_name",
    "Display Order": "sku_display_order",
    "Active": "sku_active_flag",
    "Description": "sku_description",
    "Size": "size_nominal",
    "Length (ft)": "length_feet",
    "Popularity Score": "popularity_score",
    "Eagle Sticks/Bundle": "eagle_sticks_per_bundle",
    "Eagle Bundles/Truck": "eagle_bundles_per_truckload",
    "Calc Sticks/Bundle": "calculated_sticks_per_bundle",
    "Actual Sticks/Truck": "actual_sticks_per_truckload",
    "Notes": "notes",
}

# IDs (sku_id, product_type_id) are internal — auto-assigned on publish from the
# current snapshot or generated sequentially for new entries. Users never see them.
#
# Eagle reference strings (eagle_sticks_per_bundle, eagle_bundles_per_truckload)
# are catalog references from Eagle's tables, e.g. "6 / 8" or "4 / 4".
# They are NOT used for calculations — they are display/reference only.
#
# calculated_sticks_per_bundle: the admin-selected bundle size used by the solver.
#   This must be manually chosen from the Eagle reference options.
#
# actual_sticks_per_truckload: the independently measured total sticks on a truck.
REQUIRED_COLUMNS = [
    "product_type_name",
    "product_type_display_order",
    "product_type_active_flag",
    "sku_display_order",
    "sku_active_flag",
    "sku_description",
    "size_nominal",
    "length_feet",
    "popularity_score",
    "eagle_sticks_per_bundle",
    "eagle_bundles_per_truckload",
    "calculated_sticks_per_bundle",
    "actual_sticks_per_truckload",
    "notes",
]


class ValidationIssue(BaseModel):
    level: str
    message: str
    row: int | None = None
    field: str | None = None


class SKURecord(BaseModel):
    model_config = ConfigDict(str_strip_whitespace=True)

    product_type_name: str = Field(min_length=1)
    product_type_display_order: int = Field(ge=1)
    product_type_active_flag: str
    sku_display_order: int = Field(ge=1)
    sku_active_flag: str
    sku_description: str = Field(min_length=1)
    size_nominal: str = Field(min_length=1)
    length_feet: int = Field(ge=1)
    popularity_score: int = Field(ge=1)
    eagle_sticks_per_bundle: str = Field(min_length=1)    # Eagle catalog reference string
    eagle_bundles_per_truckload: str = Field(min_length=1) # Eagle catalog reference string
    calculated_sticks_per_bundle: int = Field(ge=1)        # Admin-selected value for solver
    actual_sticks_per_truckload: int = Field(ge=1)         # Independently measured
    notes: str = ""

    @field_validator("product_type_active_flag", "sku_active_flag")
    @classmethod
    def _validate_flag(cls, value: str) -> str:
        normalized = value.strip().upper()
        if normalized not in {"Y", "N"}:
            raise ValueError("must be Y or N")
        return normalized

    @field_validator("product_type_name", "sku_description", "size_nominal")
    @classmethod
    def _validate_non_empty_text(cls, value: str) -> str:
        normalized = value.strip()
        if not normalized:
            raise ValueError("must not be empty")
        return normalized

    @field_validator("eagle_sticks_per_bundle", "eagle_bundles_per_truckload")
    @classmethod
    def _validate_eagle_ref(cls, value: str) -> str:
        normalized = value.strip()
        if not normalized:
            raise ValueError("must not be empty")
        return normalized

    @field_validator("notes", mode="before")
    @classmethod
    def _normalize_notes(cls, value: Any) -> str:
        if value is None:
            return ""
        return str(value).strip()


def _sku_key(sku_description: str, size_nominal: str) -> str:
    """Composite key uniquely identifying a SKU within the dataset.

    Uses JSON encoding to avoid false collisions when either field contains
    the separator string.
    """
    return json.dumps([sku_description, size_nominal], ensure_ascii=False)


class ValidationResult(BaseModel):
    is_valid: bool
    errors: list[ValidationIssue]
    warnings: list[ValidationIssue]
    product_types: dict[str, dict[str, Any]]
    counts: dict[str, int]


def _coerce_cell_value(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    if isinstance(value, float) and value.is_integer():
        return int(value)
    return value


def _load_workbook_bytes(source: str | Path | bytes | BinaryIO) -> bytes:
    if isinstance(source, bytes):
        return source
    if isinstance(source, (str, Path)):
        return Path(source).read_bytes()
    return source.read()


def _read_rows(source: str | Path | bytes | BinaryIO) -> tuple[list[str], list[dict[str, Any]]]:
    try:
        workbook = load_workbook(filename=BytesIO(_load_workbook_bytes(source)), data_only=True)
    except Exception as exc:
        raise ValueError(f"Could not open workbook: {exc}") from exc
    if SHEET_NAME not in workbook.sheetnames:
        raise ValueError(f"Workbook must contain a '{SHEET_NAME}' sheet")

    sheet = workbook[SHEET_NAME]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise ValueError("Workbook is empty")

    header = [_coerce_cell_value(value) for value in rows[0]]
    if not any(header):
        raise ValueError("Workbook header row is empty")

    normalized_header = [str(value).strip() for value in header]
    data_rows: list[dict[str, Any]] = []
    for row_index, row in enumerate(rows[1:], start=2):
        values = [_coerce_cell_value(value) for value in row]
        if not any(value != "" for value in values):
            continue
        row_map = {
            normalized_header[idx]: values[idx] if idx < len(values) else ""
            for idx in range(len(normalized_header))
            if normalized_header[idx]
        }
        row_map["_row_number"] = row_index
        data_rows.append(row_map)

    return normalized_header, data_rows


def _find_header_row(rows: list[tuple], expected_first_col: str) -> int:
    """Return the 0-based index of the row whose first cell matches expected_first_col.

    This handles both old format (header at index 0) and new format (title at index 0,
    header at index 1) without hardcoding row positions.
    """
    for i, row in enumerate(rows):
        first = str(_coerce_cell_value(row[0]) if row else "").strip()
        if first == expected_first_col:
            return i
    raise ValueError(f"Could not find column header '{expected_first_col}' in sheet")


def _read_rows_two_sheet(workbook: Any) -> tuple[list[str], list[dict[str, Any]]]:
    """Read the two-sheet export format: merge 'Product Types' and 'SKUs' sheets."""
    # ── Product Types sheet ───────────────────────────────────────────────────
    pt_sheet = workbook[SHEET_NAME_PRODUCT_TYPES]
    pt_raw = list(pt_sheet.iter_rows(values_only=True))

    pt_header_idx = _find_header_row(pt_raw, "Name")
    pt_header = [str(_coerce_cell_value(v)).strip() for v in pt_raw[pt_header_idx]]

    product_type_data: dict[str, dict[str, Any]] = {}
    for row in pt_raw[pt_header_idx + 1:]:
        values = [_coerce_cell_value(v) for v in row]
        if not any(v != "" for v in values):
            continue
        row_map: dict[str, Any] = {
            _TWO_SHEET_PT_MAP.get(pt_header[i], pt_header[i]): values[i] if i < len(values) else ""
            for i in range(len(pt_header))
            if pt_header[i]
        }
        pt_name = str(row_map.get("product_type_name", "")).strip()
        if pt_name:
            product_type_data[pt_name] = {
                "product_type_display_order": row_map.get("product_type_display_order", ""),
                "product_type_active_flag": row_map.get("product_type_active_flag", ""),
            }

    # ── SKUs sheet ────────────────────────────────────────────────────────────
    if SHEET_NAME not in workbook.sheetnames:
        raise ValueError(f"Workbook must contain a '{SHEET_NAME}' sheet")

    sku_sheet = workbook[SHEET_NAME]
    sku_raw = list(sku_sheet.iter_rows(values_only=True))

    sku_header_idx = _find_header_row(sku_raw, "Product Type")
    sku_header = [str(_coerce_cell_value(v)).strip() for v in sku_raw[sku_header_idx]]

    data_rows: list[dict[str, Any]] = []
    for row_index, row in enumerate(sku_raw[sku_header_idx + 1:], start=sku_header_idx + 2):
        values = [_coerce_cell_value(v) for v in row]
        if not any(v != "" for v in values):
            continue
        row_map = {
            _TWO_SHEET_SKU_MAP.get(sku_header[i], sku_header[i]): values[i] if i < len(values) else ""
            for i in range(len(sku_header))
            if sku_header[i]
        }
        pt_name = str(row_map.get("product_type_name", "")).strip()
        pt_info = product_type_data.get(pt_name, {})
        row_map["product_type_display_order"] = pt_info.get("product_type_display_order", "")
        row_map["product_type_active_flag"] = pt_info.get("product_type_active_flag", "")
        row_map.setdefault("notes", "")
        row_map["_row_number"] = row_index
        data_rows.append(row_map)

    # Return REQUIRED_COLUMNS as the header — all 14 fields are present after the merge
    return REQUIRED_COLUMNS, data_rows


def _validate_columns(header: list[str]) -> list[ValidationIssue]:
    issues: list[ValidationIssue] = []
    missing = [column for column in REQUIRED_COLUMNS if column not in header]
    extra = [column for column in header if column and column not in REQUIRED_COLUMNS]

    for column in missing:
        issues.append(
            ValidationIssue(level="error", field=column, message=f"Missing required column: {column}")
        )
    for column in extra:
        issues.append(
            ValidationIssue(level="warning", field=column, message=f"Unexpected column ignored: {column}")
        )
    return issues


def _build_issue_from_validation_error(row_number: int, error: dict[str, Any]) -> ValidationIssue:
    location = error.get("loc", ())
    field = location[0] if location else None
    return ValidationIssue(
        level="error",
        row=row_number,
        field=str(field) if field is not None else None,
        message=error.get("msg", "Invalid value"),
    )


def _validate_rows(
    data_rows: list[dict[str, Any]],
) -> tuple[list[tuple[int, SKURecord]], list[ValidationIssue]]:
    valid_rows: list[tuple[int, SKURecord]] = []
    issues: list[ValidationIssue] = []

    for row in data_rows:
        row_number = int(row.get("_row_number", 0))
        payload = {column: row.get(column, "") for column in REQUIRED_COLUMNS}
        try:
            valid_rows.append((row_number, SKURecord(**payload)))
        except ValidationError as exc:
            for error in exc.errors():
                issues.append(_build_issue_from_validation_error(row_number, error))

    return valid_rows, issues


def _enforce_cross_row_rules(
    records: list[tuple[int, SKURecord]],
) -> tuple[list[ValidationIssue], list[ValidationIssue]]:
    errors: list[ValidationIssue] = []
    warnings: list[ValidationIssue] = []

    seen_composite: dict[str, int] = {}
    product_type_state: dict[str, dict[str, Any]] = {}

    for row_number, record in records:
        composite = json.dumps([record.product_type_name, record.sku_description, record.size_nominal], ensure_ascii=False)
        if composite in seen_composite:
            first_row = seen_composite[composite]
            errors.append(
                ValidationIssue(
                    level="error",
                    row=row_number,
                    field="sku_description",
                    message=(
                        f"Duplicate SKU: '{record.sku_description}' ({record.size_nominal}) "
                        f"in '{record.product_type_name}' (first seen at row {first_row})"
                    ),
                )
            )
        else:
            seen_composite[composite] = row_number

        existing = product_type_state.get(record.product_type_name)
        current = {
            "product_type_display_order": record.product_type_display_order,
            "product_type_active_flag": record.product_type_active_flag,
        }
        if existing is None:
            product_type_state[record.product_type_name] = current
        elif existing != current:
            errors.append(
                ValidationIssue(
                    level="error",
                    row=row_number,
                    field="product_type_name",
                    message=(
                        f"Product type '{record.product_type_name}' has inconsistent "
                        f"display_order or active_flag values across rows"
                    ),
                )
            )

    grouped: dict[str, list[tuple[int, SKURecord]]] = {}
    for row_number, record in records:
        grouped.setdefault(record.product_type_name, []).append((row_number, record))

    for product_type_name, sku_records in grouped.items():
        if all(record.sku_active_flag == "N" for _, record in sku_records):
            first_row_number, exemplar = sku_records[0]
            if exemplar.product_type_active_flag == "Y":
                warnings.append(
                    ValidationIssue(
                        level="warning",
                        row=first_row_number,
                        field="product_type_active_flag",
                        message=(
                            f"Product type '{product_type_name}' will be deactivated "
                            f"because all its SKUs are inactive"
                        ),
                    )
                )

    return errors, warnings


def _build_normalized_structure(records: list[tuple[int, SKURecord]]) -> dict[str, dict[str, Any]]:
    """Build product_types dict keyed by product_type_name.
    Each product type's skus dict is keyed by _sku_key(desc, size)."""
    product_types: dict[str, dict[str, Any]] = {}

    for _, record in records:
        product_type = product_types.setdefault(
            record.product_type_name,
            {
                "product_type_name": record.product_type_name,
                "product_type_display_order": record.product_type_display_order,
                "product_type_active_flag": record.product_type_active_flag,
                "skus": {},
            },
        )

        product_type["skus"][_sku_key(record.sku_description, record.size_nominal)] = {
            "sku_description": record.sku_description,
            "size_nominal": record.size_nominal,
            "sku_display_order": record.sku_display_order,
            "sku_active_flag": record.sku_active_flag,
            "length_feet": record.length_feet,
            "popularity_score": record.popularity_score,
            "eagle_sticks_per_bundle": record.eagle_sticks_per_bundle,
            "eagle_bundles_per_truckload": record.eagle_bundles_per_truckload,
            "calculated_sticks_per_bundle": record.calculated_sticks_per_bundle,
            "actual_sticks_per_truckload": record.actual_sticks_per_truckload,
            "notes": record.notes,
        }

    for product_type in product_types.values():
        if all(sku["sku_active_flag"] == "N" for sku in product_type["skus"].values()):
            product_type["product_type_active_flag"] = "N"

    return product_types


def import_workbook(source: str | Path | bytes | BinaryIO) -> ValidationResult:
    try:
        raw_bytes = _load_workbook_bytes(source)
        try:
            workbook = load_workbook(filename=BytesIO(raw_bytes), data_only=True)
        except Exception as exc:
            raise ValueError(f"Could not open workbook: {exc}") from exc

        if SHEET_NAME_PRODUCT_TYPES in workbook.sheetnames:
            header, data_rows = _read_rows_two_sheet(workbook)
        else:
            header, data_rows = _read_rows(raw_bytes)
    except ValueError as exc:
        return ValidationResult(
            is_valid=False,
            errors=[ValidationIssue(level="error", message=str(exc))],
            warnings=[],
            product_types={},
            counts={"rows": 0, "product_types": 0, "skus": 0},
        )

    column_issues = _validate_columns(header)

    if any(issue.level == "error" for issue in column_issues):
        return ValidationResult(
            is_valid=False,
            errors=[issue for issue in column_issues if issue.level == "error"],
            warnings=[issue for issue in column_issues if issue.level == "warning"],
            product_types={},
            counts={"rows": len(data_rows), "product_types": 0, "skus": 0},
        )

    valid_rows, row_errors = _validate_rows(data_rows)
    cross_row_errors, cross_row_warnings = _enforce_cross_row_rules(valid_rows)

    errors = [issue for issue in column_issues if issue.level == "error"] + row_errors + cross_row_errors
    warnings = [issue for issue in column_issues if issue.level == "warning"] + cross_row_warnings

    if errors:
        return ValidationResult(
            is_valid=False,
            errors=errors,
            warnings=warnings,
            product_types={},
            counts={"rows": len(data_rows), "product_types": 0, "skus": 0},
        )

    product_types = _build_normalized_structure(valid_rows)
    return ValidationResult(
        is_valid=True,
        errors=[],
        warnings=warnings,
        product_types=product_types,
        counts={
            "rows": len(data_rows),
            "product_types": len(product_types),
            "skus": len(valid_rows),
        },
    )
