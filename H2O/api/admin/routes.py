from __future__ import annotations

import hashlib
import json
from io import BytesIO
from datetime import datetime
from typing import Any
from pathlib import Path

from fastapi import APIRouter, File, Form, Header, HTTPException, UploadFile
from fastapi.responses import JSONResponse, StreamingResponse
from loguru import logger
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from auth import require_admin, require_review
from data import load_snapshot
import runtime
from admin.data_import import REQUIRED_COLUMNS, ValidationIssue, ValidationResult, import_workbook, _sku_key

router = APIRouter(prefix="/admin/data", tags=["admin-data"])

MAX_UPLOAD_BYTES = 10 * 1024 * 1024  # 10 MB


def _require_xlsx(upload: UploadFile) -> None:
    if not upload.filename or not upload.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Only .xlsx uploads are supported")


def _check_file_size(data: bytes) -> None:
    if len(data) > MAX_UPLOAD_BYTES:
        raise HTTPException(status_code=413, detail="File too large (max 10 MB)")


def _coerce_int(value: Any) -> int:
    if isinstance(value, int):
        return value
    return int(str(value).replace(",", "").strip())


def _coerce_length_feet(value: Any) -> int:
    text = str(value).strip().replace("'", "").replace("ft", "").strip()
    return int(text)


def _normalize_current_snapshot(snapshot: dict[str, Any]) -> dict[str, dict[str, Any]]:
    """Convert the legacy JSON snapshot to the same format produced by import_workbook.

    Product types are keyed by product_type_name.
    SKUs are keyed by _sku_key(sku_description, size_nominal).
    """
    product_types: dict[str, dict[str, Any]] = {}

    for display_order, product_type in enumerate(snapshot.get("productTypes", []), start=1):
        pt_name = product_type.get("productType", "")
        if not pt_name:
            continue

        if pt_name not in product_types:
            product_types[pt_name] = {
                "product_type_name": pt_name,
                "product_type_display_order": display_order,
                "product_type_active_flag": "Y",
                "skus": {},
            }

        for _sku_id, sku in product_type.get("skus", {}).items():
            sku_desc = sku.get("SKU", "")
            size = sku.get("size", "")
            # Eagle reference strings — stored as-is, never parsed as numbers
            eagle_sticks = str(sku.get("eagleSticksPerBundle", "")).strip()
            eagle_bundles = str(sku.get("eagleBundlesPerTruckLoad", "")).strip()
            # calculatedSticksPerBundle is the admin-selected value, always an integer
            calc_sticks = int(sku.get("calculatedSticksPerBundle", 0))
            actual_sticks = _coerce_int(
                sku.get("actualSticksPerTruckLoad", sku.get("eagleSticksPerTruckload", 0))
            )

            product_types[pt_name]["skus"][_sku_key(sku_desc, size)] = {
                "sku_description": sku_desc,
                "size_nominal": size,
                "sku_display_order": int(sku.get("displayOrder", 0)),
                "sku_active_flag": "Y",
                "length_feet": _coerce_length_feet(sku.get("length", "0")),
                "popularity_score": int(sku.get("popularityScore", 1)),
                "eagle_sticks_per_bundle": eagle_sticks,
                "eagle_bundles_per_truckload": eagle_bundles,
                "calculated_sticks_per_bundle": calc_sticks,
                "actual_sticks_per_truckload": actual_sticks,
                "notes": "",
            }

    return product_types


def _response_from_validation(validation: ValidationResult, status_code: int) -> JSONResponse:
    return JSONResponse(status_code=status_code, content=validation.model_dump())


def _build_diff_summary(
    current_product_types: dict[str, dict[str, Any]],
    uploaded_product_types: dict[str, dict[str, Any]],
) -> dict[str, Any]:
    def _flatten(product_types: dict[str, dict[str, Any]]) -> dict[str, dict[str, Any]]:
        result: dict[str, dict[str, Any]] = {}
        for pt_name, pt in product_types.items():
            for sk, sku in pt["skus"].items():
                display_key = f"{sku['sku_description']} ({sku['size_nominal']})"
                result[display_key] = {**sku, "_pt_name": pt_name}
        return result

    current_flat = _flatten(current_product_types)
    uploaded_flat = _flatten(uploaded_product_types)

    added_skus = sorted(set(uploaded_flat) - set(current_flat))
    removed_skus = sorted(set(current_flat) - set(uploaded_flat))
    shared_skus = sorted(set(current_flat) & set(uploaded_flat))

    fields_to_compare = [
        "sku_display_order",
        "sku_active_flag",
        "length_feet",
        "popularity_score",
        "eagle_sticks_per_bundle",
        "eagle_bundles_per_truckload",
        "actual_sticks_per_truckload",
        "notes",
    ]

    changed_skus: dict[str, list[dict[str, Any]]] = {}
    active_flag_changes: list[dict[str, Any]] = []

    for key in shared_skus:
        before = current_flat[key]
        after = uploaded_flat[key]
        changes: list[dict[str, Any]] = []
        for field in fields_to_compare:
            if before.get(field) != after.get(field):
                change = {"field": field, "before": before.get(field), "after": after.get(field)}
                if field == "sku_active_flag":
                    change["highlight"] = "active_flag"
                    active_flag_changes.append(
                        {"sku": key, "before": before.get(field), "after": after.get(field)}
                    )
                changes.append(change)
        if changes:
            changed_skus[key] = changes

    return {
        "counts": {
            "current": {
                "product_types": len(current_product_types),
                "skus": len(current_flat),
            },
            "uploaded": {
                "product_types": len(uploaded_product_types),
                "skus": len(uploaded_flat),
            },
        },
        "added_skus": added_skus,
        "removed_skus": removed_skus,
        "changed_skus": changed_skus,
        "active_flag_changes": active_flag_changes,
    }


# ── Excel styling constants ───────────────────────────────────────────────────
_TITLE_FILL = PatternFill("solid", fgColor="1F4E79")
_TITLE_FONT = Font(bold=True, color="FFFFFF", name="Calibri", size=14)
_TITLE_ALIGN = Alignment(horizontal="center", vertical="center")

_HEADER_FILL = PatternFill("solid", fgColor="2E75B6")
_HEADER_FONT = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center")

_ALT_ROW_FILL = PatternFill("solid", fgColor="EEF3FA")
_Y_FILL = PatternFill("solid", fgColor="C6EFCE")   # light green
_N_FILL = PatternFill("solid", fgColor="FFC7CE")   # light red
_CENTER = Alignment(horizontal="center", vertical="center")
_LEFT = Alignment(horizontal="left", vertical="center")


def _add_title_row(sheet: Any, title: str, num_cols: int) -> None:
    """Merge row 1 across all columns and apply title styling."""
    last_col = get_column_letter(num_cols)
    sheet.merge_cells(f"A1:{last_col}1")
    cell = sheet["A1"]
    cell.value = title
    cell.fill = _TITLE_FILL
    cell.font = _TITLE_FONT
    cell.alignment = _TITLE_ALIGN
    sheet.row_dimensions[1].height = 30


def _style_sheet(sheet: Any, active_flag_cols: list[str]) -> None:
    """Style column header row (row 2), freeze at row 3, autofilter, row colors."""
    # Column header row (row 2, below the title)
    for cell in sheet[2]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _HEADER_ALIGN
    sheet.row_dimensions[2].height = 22
    sheet.freeze_panes = "A3"
    last_col = get_column_letter(sheet.max_column)
    sheet.auto_filter.ref = f"A2:{last_col}2"

    # Data rows: alternating background + active flag coloring
    active_col_set = set(active_flag_cols)
    for row_idx, row in enumerate(sheet.iter_rows(min_row=3, max_row=sheet.max_row), start=0):
        alt = row_idx % 2 == 1
        for cell in row:
            if cell.value is None:
                continue
            col = get_column_letter(cell.column)
            if col in active_col_set:
                val = str(cell.value).upper()
                cell.fill = _Y_FILL if val == "Y" else _N_FILL
                cell.alignment = _CENTER
            elif alt:
                cell.fill = _ALT_ROW_FILL


def _build_current_excel_workbook(product_types: dict[str, dict[str, Any]]) -> bytes:
    """Export current data as a two-sheet Excel workbook.

    Sheet 1 — "Product Types": Name | Display Order | Active
    Sheet 2 — "SKUs": Product Type (dropdown) | Display Order | Active | Description |
              Size | Length (ft) | Popularity Score | Eagle Sticks/Bundle |
              Eagle Bundles/Truck | Calc Sticks/Bundle | Actual Sticks/Truck | Notes
    """
    workbook = Workbook()

    sorted_pts = sorted(
        product_types.values(),
        key=lambda pt: (int(pt["product_type_display_order"]), pt["product_type_name"].lower()),
    )

    # ── Sheet 1: Product Types ────────────────────────────────────────────────
    _active = workbook.active
    assert _active is not None
    pt_sheet = _active
    pt_sheet.title = "Product Types"
    pt_sheet.append([""])                                    # row 1 — title (merged below)
    pt_sheet.append(["Name", "Display Order", "Active"])    # row 2 — column headers

    for pt in sorted_pts:
        pt_sheet.append([
            pt["product_type_name"],
            int(pt["product_type_display_order"]),
            pt["product_type_active_flag"],
        ])

    num_pts = len(sorted_pts)
    pt_max = max(num_pts + 3, 100)   # +2 for title+header rows

    # Active flag: Y/N dropdown (column C) — data starts at row 3
    yn_pt = DataValidation(type="list", formula1='"Y,N"', allow_blank=False)
    yn_pt.sqref = f"C3:C{pt_max}"
    pt_sheet.add_data_validation(yn_pt)

    # Display order: whole number ≥ 1 (column B)
    int_pt = DataValidation(type="whole", operator="greaterThanOrEqual", formula1="1", allow_blank=False)
    int_pt.sqref = f"B3:B{pt_max}"
    pt_sheet.add_data_validation(int_pt)

    _add_title_row(pt_sheet, "H2O — Product Types", num_cols=3)
    pt_sheet.column_dimensions["A"].width = 40
    pt_sheet.column_dimensions["B"].width = 14
    pt_sheet.column_dimensions["C"].width = 8
    _style_sheet(pt_sheet, active_flag_cols=["C"])

    # ── Sheet 2: SKUs ────────────────────────────────────────────────────────
    sku_sheet = workbook.create_sheet("SKUs")
    sku_sheet.append([""])                                   # row 1 — title (merged below)
    sku_sheet.append([                                       # row 2 — column headers
        "Product Type", "Display Order", "Active", "Description", "Size",
        "Length (ft)", "Popularity Score", "Eagle Sticks/Bundle", "Eagle Bundles/Truck",
        "Calc Sticks/Bundle", "Actual Sticks/Truck", "Notes",
    ])

    total_skus = sum(len(pt["skus"]) for pt in sorted_pts)
    sku_max = max(total_skus + 3, 200)   # +2 for title+header rows

    for pt in sorted_pts:
        sorted_skus = sorted(
            pt["skus"].values(),
            key=lambda s: (int(s["sku_display_order"]), s["sku_description"].lower()),
        )
        for sku in sorted_skus:
            sku_sheet.append([
                pt["product_type_name"],
                int(sku["sku_display_order"]),
                sku["sku_active_flag"],
                sku["sku_description"],
                sku["size_nominal"],
                int(sku["length_feet"]),
                int(sku["popularity_score"]),
                str(sku["eagle_sticks_per_bundle"]),
                str(sku["eagle_bundles_per_truckload"]),
                int(sku["calculated_sticks_per_bundle"]),
                int(sku["actual_sticks_per_truckload"]),
                sku.get("notes", ""),
            ])

    # Product Type: dropdown from Product Types data rows (A3 onwards after title)
    pt_dropdown = DataValidation(
        type="list",
        formula1=f"'Product Types'!$A$3:$A${num_pts + 2}",
        allow_blank=False,
    )
    pt_dropdown.sqref = f"A3:A{sku_max}"
    sku_sheet.add_data_validation(pt_dropdown)

    # Active flag: Y/N dropdown (column C) — data starts at row 3
    yn_sku = DataValidation(type="list", formula1='"Y,N"', allow_blank=False)
    yn_sku.sqref = f"C3:C{sku_max}"
    sku_sheet.add_data_validation(yn_sku)

    # Integer ≥ 1: Display Order (B), Length (F), Popularity Score (G),
    #               Calc Sticks/Bundle (J), Actual Sticks/Truck (K)
    for col in ("B", "F", "G", "J", "K"):
        val = DataValidation(type="whole", operator="greaterThanOrEqual", formula1="1", allow_blank=False)
        val.sqref = f"{col}3:{col}{sku_max}"
        sku_sheet.add_data_validation(val)

    _add_title_row(sku_sheet, "H2O — SKUs", num_cols=12)
    col_widths = [40, 14, 8, 50, 8, 12, 16, 22, 22, 18, 18, 30]
    for i, width in enumerate(col_widths, start=1):
        sku_sheet.column_dimensions[get_column_letter(i)].width = width

    _style_sheet(sku_sheet, active_flag_cols=["C"])

    # Per-column alignment for SKU data rows (left for text, center for numbers/flags)
    # Columns: A=left, B=center, C=handled by _style_sheet, D=left, E=center,
    #          F=center, G=center, H=left, I=left, J=center, K=center, L=left
    _left_cols = {"A", "D", "H", "I", "L"}
    _center_cols = {"B", "E", "F", "G", "J", "K"}
    for row in sku_sheet.iter_rows(min_row=3, max_row=sku_sheet.max_row):
        for cell in row:
            if cell.value is None:
                continue
            col = get_column_letter(cell.column)
            if col in _left_cols:
                cell.alignment = _LEFT
            elif col in _center_cols:
                cell.alignment = _CENTER

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _assign_ids_from_snapshot(
    imported_product_types: dict[str, dict[str, Any]],
    current_snapshot: dict[str, Any],
) -> dict[str, dict[str, Any]]:
    """Assign stable IDs to imported data by matching against the current snapshot.

    Existing product_type_names and (pt_name, sku_description, size_nominal) tuples
    reuse their existing IDs. New entries get the next sequential ID.
    """
    pt_name_to_id: dict[str, str] = {}
    sku_composite_to_id: dict[str, str] = {}
    existing_pt_ids: set[str] = set()
    existing_sku_ids: set[str] = set()

    for pt in current_snapshot.get("productTypes", []):
        pt_id = pt["productTypeId"]
        pt_name = pt.get("productType", "")
        pt_name_to_id[pt_name] = pt_id
        existing_pt_ids.add(pt_id)
        for sku_id, sku in pt.get("skus", {}).items():
            existing_sku_ids.add(sku_id)
            composite = json.dumps([pt_name, sku.get("SKU", ""), sku.get("size", "")], ensure_ascii=False)
            sku_composite_to_id[composite] = sku_id

    def _next_pt_id() -> str:
        nums = [
            int(pid[len("productTypeId_"):])
            for pid in existing_pt_ids
            if pid.startswith("productTypeId_") and pid[len("productTypeId_"):].isdigit()
        ]
        n = max(nums, default=0) + 1
        while f"productTypeId_{n}" in existing_pt_ids:
            n += 1
        new_id = f"productTypeId_{n}"
        existing_pt_ids.add(new_id)
        return new_id

    def _next_sku_id() -> str:
        nums = [
            int(sid[len("skuId_"):])
            for sid in existing_sku_ids
            if sid.startswith("skuId_") and sid[len("skuId_"):].isdigit()
        ]
        n = max(nums, default=0) + 1
        while f"skuId_{n}" in existing_sku_ids:
            n += 1
        new_id = f"skuId_{n}"
        existing_sku_ids.add(new_id)
        return new_id

    result: dict[str, dict[str, Any]] = {}

    for pt_name, pt_data in imported_product_types.items():
        pt_id = pt_name_to_id.get(pt_name)
        if pt_id is None:
            pt_id = _next_pt_id()
            pt_name_to_id[pt_name] = pt_id

        id_assigned_pt: dict[str, Any] = {
            "product_type_id": pt_id,
            "product_type_name": pt_data["product_type_name"],
            "product_type_display_order": pt_data["product_type_display_order"],
            "product_type_active_flag": pt_data["product_type_active_flag"],
            "skus": {},
        }

        for _sk, sku_data in pt_data["skus"].items():
            composite = json.dumps([pt_name, sku_data["sku_description"], sku_data["size_nominal"]], ensure_ascii=False)
            sku_id = sku_composite_to_id.get(composite)
            if sku_id is None:
                sku_id = _next_sku_id()
                sku_composite_to_id[composite] = sku_id

            id_assigned_pt["skus"][sku_id] = {
                "sku_id": sku_id,
                "product_type_id": pt_id,
                **sku_data,
            }

        result[pt_id] = id_assigned_pt

    return result


def _serialize_legacy_snapshot(product_types: dict[str, dict[str, Any]]) -> dict[str, Any]:
    """Convert ID-assigned product_types back to the legacy JSON snapshot format."""
    snapshot_product_types: list[dict[str, Any]] = []

    active_pts = sorted(
        (pt for pt in product_types.values() if pt["product_type_active_flag"] == "Y"),
        key=lambda pt: (int(pt["product_type_display_order"]), pt["product_type_name"].lower()),
    )

    for pt in active_pts:
        active_skus = sorted(
            (sku for sku in pt["skus"].values() if sku["sku_active_flag"] == "Y"),
            key=lambda s: (int(s["sku_display_order"]), int(s["popularity_score"]), s["sku_description"].lower()),
        )
        if not active_skus:
            continue

        legacy_skus: dict[str, dict[str, Any]] = {}
        for sku in active_skus:
            actual = int(sku["actual_sticks_per_truckload"])
            calculated_bundle_size = int(sku["calculated_sticks_per_bundle"])
            legacy_skus[sku["sku_id"]] = {
                "skuId": sku["sku_id"],
                "productTypeId": sku["product_type_id"],
                "displayOrder": int(sku["sku_display_order"]),
                "popularityScore": int(sku["popularity_score"]),
                "productType": pt["product_type_name"],
                "SKU": sku["sku_description"],
                "eagleSticksPerBundle": str(sku["eagle_sticks_per_bundle"]),
                "eagleBundlesPerTruckLoad": str(sku["eagle_bundles_per_truckload"]),
                "eagleSticksPerTruckload": actual,
                "actualSticksPerTruckLoad": actual,
                "length": f"{int(sku['length_feet'])}'",
                "calculatedSticksPerBundle": calculated_bundle_size,
                "size": sku["size_nominal"],
            }

        snapshot_product_types.append({
            "productTypeId": pt["product_type_id"],
            "productType": pt["product_type_name"],
            "skus": legacy_skus,
        })

    return {
        "date": datetime.now().strftime("%Y-%m-%d_%H-%M-%S"),
        "productTypes": snapshot_product_types,
    }


def _write_active_snapshot(snapshot: dict[str, Any]) -> tuple[str, str]:
    version_id = datetime.now().strftime("%Y%m%d_%H%M%S")
    json_text = json.dumps(snapshot, indent=4)
    json_hash = hashlib.sha256(json_text.encode("utf-8")).hexdigest()

    runtime.VERSIONS_DIR.mkdir(parents=True, exist_ok=True)
    versioned_path = runtime.VERSIONS_DIR / f"{version_id}_packing_data.json"
    versioned_path.write_text(json_text, encoding="utf-8")

    return version_id, json_hash


def _write_manifest(
    version_id: str,
    source_filename: str,
    change_reason: str,
    validation: ValidationResult,
    json_hash: str,
) -> Path:
    manifest_path = runtime.VERSIONS_DIR / f"{version_id}_manifest.json"
    manifest = {
        "version_id": version_id,
        "published_at": datetime.now().isoformat(),
        "source_filename": source_filename,
        "change_reason": change_reason,
        "counts": validation.counts,
        "sha256": json_hash,
        "warnings": [w.model_dump() for w in validation.warnings],
    }
    manifest_path.write_text(json.dumps(manifest, indent=4), encoding="utf-8")
    return manifest_path


def _activate_snapshot(snapshot: dict[str, Any]) -> None:
    runtime.SNAPSHOT_PATH.parent.mkdir(parents=True, exist_ok=True)
    temp_path = runtime.SNAPSHOT_PATH.with_suffix(".json.tmp")
    temp_path.write_text(json.dumps(snapshot, indent=4), encoding="utf-8")
    temp_path.replace(runtime.SNAPSHOT_PATH)
    runtime.reload_runtime_snapshot(runtime.SNAPSHOT_PATH)


# ---------------------------------------------------------------------------
# Endpoints
# ---------------------------------------------------------------------------

@router.get("/current")
async def get_current_data(
    authorization: str | None = Header(default=None),
) -> dict[str, Any]:
    """Return the current active packing data in a browser-friendly format."""
    require_review(authorization)
    current_snapshot = load_snapshot(runtime.SNAPSHOT_PATH)
    product_types = _normalize_current_snapshot(current_snapshot)

    sorted_pts = sorted(
        product_types.values(),
        key=lambda pt: (int(pt["product_type_display_order"]), pt["product_type_name"].lower()),
    )
    for pt in sorted_pts:
        pt["skus"] = dict(
            sorted(
                pt["skus"].items(),
                key=lambda kv: (int(kv[1]["sku_display_order"]), kv[1]["sku_description"].lower()),
            )
        )

    return {
        "product_types": {pt["product_type_name"]: pt for pt in sorted_pts},
        "counts": {
            "product_types": len(sorted_pts),
            "skus": sum(len(pt["skus"]) for pt in sorted_pts),
        },
    }


@router.get("/versions")
async def list_versions(
    authorization: str | None = Header(default=None),
) -> dict[str, Any]:
    """Return the last 20 version manifests, newest first."""
    require_admin(authorization)
    if not runtime.VERSIONS_DIR.exists():
        return {"versions": []}

    manifests = sorted(
        runtime.VERSIONS_DIR.glob("*_manifest.json"),
        key=lambda p: p.name,
        reverse=True,
    )[:20]

    versions = []
    for mp in manifests:
        try:
            versions.append(json.loads(mp.read_text(encoding="utf-8")))
        except Exception as exc:
            logger.warning("Failed to parse manifest {}: {}", mp.name, exc)

    return {"versions": versions}


@router.get("/current_excel")
async def download_current_excel(
    authorization: str | None = Header(default=None),
) -> StreamingResponse:
    require_review(authorization)
    current_snapshot = load_snapshot(runtime.SNAPSHOT_PATH)
    product_types = _normalize_current_snapshot(current_snapshot)
    workbook_bytes = _build_current_excel_workbook(product_types)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"packing_data_{timestamp}.xlsx"
    headers = {"Content-Disposition": f'attachment; filename="{filename}"'}
    return StreamingResponse(
        BytesIO(workbook_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


@router.get("/current_json")
async def download_current_json(
    authorization: str | None = Header(default=None),
) -> StreamingResponse:
    require_review(authorization)
    snapshot = json.loads(runtime.SNAPSHOT_PATH.read_bytes())
    js_content = "const packingData = " + json.dumps(snapshot, indent=4) + ";"
    headers = {"Content-Disposition": 'attachment; filename="load_packing_data.js"'}
    return StreamingResponse(
        BytesIO(js_content.encode("utf-8")),
        media_type="application/javascript",
        headers=headers,
    )


@router.post("/validate")
async def validate_workbook(
    file: UploadFile = File(...),
    authorization: str | None = Header(default=None),
) -> ValidationResult:
    require_review(authorization)
    _require_xlsx(file)
    file_bytes = await file.read()
    _check_file_size(file_bytes)
    return import_workbook(file_bytes)


@router.post("/preview")
async def preview_workbook(
    file: UploadFile = File(...),
    authorization: str | None = Header(default=None),
) -> dict[str, Any]:
    require_review(authorization)
    _require_xlsx(file)
    file_bytes = await file.read()
    _check_file_size(file_bytes)
    validation = import_workbook(file_bytes)
    if not validation.is_valid:
        return _response_from_validation(validation, 400)

    current_snapshot = load_snapshot(runtime.SNAPSHOT_PATH)
    current_product_types = _normalize_current_snapshot(current_snapshot)
    diff = _build_diff_summary(current_product_types, validation.product_types)
    return {"validation": validation.model_dump(), "diff": diff}


@router.post("/publish")
async def publish_workbook(
    file: UploadFile = File(...),
    change_reason: str = Form(...),
    authorization: str | None = Header(default=None),
) -> dict[str, Any]:
    require_admin(authorization)
    _require_xlsx(file)
    if not change_reason or not change_reason.strip():
        raise HTTPException(status_code=400, detail="change_reason is required")

    file_bytes = await file.read()
    _check_file_size(file_bytes)
    validation = import_workbook(file_bytes)
    if not validation.is_valid:
        return _response_from_validation(validation, 400)

    current_snapshot = load_snapshot(runtime.SNAPSHOT_PATH)
    id_assigned = _assign_ids_from_snapshot(validation.product_types, current_snapshot)
    snapshot = _serialize_legacy_snapshot(id_assigned)
    version_id, json_hash = _write_active_snapshot(snapshot)
    _write_manifest(
        version_id=version_id,
        source_filename=file.filename or "upload.xlsx",
        change_reason=change_reason.strip(),
        validation=validation,
        json_hash=json_hash,
    )

    # Capture the old snapshot content so we can roll back if activation fails.
    old_snapshot_bytes: bytes | None = None
    if runtime.SNAPSHOT_PATH.exists():
        old_snapshot_bytes = runtime.SNAPSHOT_PATH.read_bytes()

    try:
        _activate_snapshot(snapshot)
    except Exception as exc:
        logger.error("Snapshot activation failed (version {}): {}", version_id, exc)
        if old_snapshot_bytes is not None:
            try:
                runtime.SNAPSHOT_PATH.write_bytes(old_snapshot_bytes)
                runtime.reload_runtime_snapshot(runtime.SNAPSHOT_PATH)
                logger.info("Rolled back to previous snapshot after activation failure")
            except Exception as rb_exc:
                logger.error("Rollback also failed: {}", rb_exc)
        raise HTTPException(status_code=500, detail=f"Snapshot activation failed: {exc}") from exc

    return {
        "ok": True,
        "version_id": version_id,
        "active_snapshot_path": str(runtime.SNAPSHOT_PATH),
        "counts": validation.counts,
    }
